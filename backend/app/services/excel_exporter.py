"""
Excel Export Service — Generates professional Excel reports.
Covers: Technical Audit, Rankings, Backlinks, Content Gaps,
Blog Ideas, Internal Links, SEO Tasks, Competitor Analysis.
"""
import io
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


# ─── Color Palette ────────────────────────────────────────────────
COLORS = {
    "header_bg":    "1E40AF",   # dark blue
    "header_fg":    "FFFFFF",
    "critical_bg":  "FEE2E2",   # light red
    "critical_fg":  "991B1B",
    "high_bg":      "FEF3C7",   # light amber
    "high_fg":      "92400E",
    "medium_bg":    "DBEAFE",   # light blue
    "medium_fg":    "1E40AF",
    "low_bg":       "D1FAE5",   # light green
    "low_fg":       "065F46",
    "alt_row":      "F8FAFC",
    "score_good":   "10B981",
    "score_mid":    "F59E0B",
    "score_bad":    "EF4444",
}


def _header_style(ws, row, col, value, width=20):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=COLORS["header_fg"], size=11)
    cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="FFFFFF"))
    ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _title_row(ws, title: str, cols: int):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14, color=COLORS["header_fg"])
    cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35


def _severity_fill(severity: str) -> PatternFill:
    mapping = {
        "critical": PatternFill("solid", fgColor=COLORS["critical_bg"]),
        "high":     PatternFill("solid", fgColor=COLORS["high_bg"]),
        "medium":   PatternFill("solid", fgColor=COLORS["medium_bg"]),
        "low":      PatternFill("solid", fgColor=COLORS["low_bg"]),
    }
    return mapping.get(severity, PatternFill("solid", fgColor="FFFFFF"))


def _score_color(score: int) -> str:
    if score >= 70: return COLORS["score_good"]
    if score >= 40: return COLORS["score_mid"]
    return COLORS["score_bad"]


# ─── Technical Audit ──────────────────────────────────────────────
def technical_audit_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Technical Audit")
    issues = data.get("issues", [])
    pages = data.get("pages", [])
    summary = data.get("summary", {})

    cols = ["#", "Issue Type", "Severity", "Page URL", "Description", "Recommendation", "Impact"]
    _title_row(ws, f"Technical SEO Audit — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    # Summary block
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Pages Crawled: {summary.get('pages_crawled', 0)}  |  Total Issues: {summary.get('total_issues', 0)}  |  Critical: {summary.get('critical_issues', 0)}  |  SEO Score: {data.get('seo_score', 0)}/100"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    row = 3
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=[4, 25, 12, 45, 40, 45, 12][i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, issue in enumerate(issues, 1):
        severity = str(issue.get("severity", "medium")).lower()
        fill = _severity_fill(severity)
        values = [
            idx,
            issue.get("issue_type", "").replace("_", " ").title(),
            severity.upper(),
            issue.get("page_url", ""),
            issue.get("description", ""),
            issue.get("recommendation", "Fix this issue to improve SEO score"),
            issue.get("impact_score", 50),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill if col in [1, 2, 3] else (PatternFill("solid", fgColor=COLORS["alt_row"]) if idx % 2 == 0 else PatternFill())
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 7:
                cell.font = Font(bold=True, color=_score_color(100 - int(val or 50)))
        ws.row_dimensions[row].height = 35
        row += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(cols))}{row-1}"


# ─── Keyword Rankings ─────────────────────────────────────────────
def rankings_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Keyword Rankings")
    rankings = data.get("rankings", [])

    cols = ["#", "Keyword", "Current Position", "Previous Position", "Change", "Search Volume", "Page URL", "Last Updated"]
    _title_row(ws, f"Keyword Rankings — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=[4, 35, 16, 16, 10, 15, 45, 15][i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, r in enumerate(rankings, 1):
        change = (r.get("previous_position", 0) or 0) - (r.get("position", 0) or 0)
        values = [idx, r.get("keyword", ""), r.get("position", "-"), r.get("previous_position", "-"), change, r.get("search_volume", 0), r.get("page_url", ""), str(r.get("updated_at", ""))[:10]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center" if col in [3,4,5,6] else "left", vertical="center")
            if col == 5 and isinstance(val, (int, float)):
                cell.font = Font(color="10B981" if val > 0 else ("EF4444" if val < 0 else "6B7280"), bold=True)
                cell.value = f"▲{val}" if val > 0 else (f"▼{abs(val)}" if val < 0 else "—")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{row-1}"


# ─── Backlink Opportunities ───────────────────────────────────────
def backlinks_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Backlink Opportunities")
    opps = data.get("opportunities", [])

    cols = ["#", "Platform", "Type", "Domain Authority", "Relevance Score", "Dofollow", "Status", "Submission URL", "Notes", "Action"]
    _title_row(ws, f"Backlink Opportunities — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 25, 18, 15, 15, 10, 15, 45, 35, 25]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, opp in enumerate(opps, 1):
        da = opp.get("domain_authority", 0)
        da_fill = PatternFill("solid", fgColor=("D1FAE5" if da >= 70 else ("FEF3C7" if da >= 40 else "FEE2E2")))
        values = [
            idx,
            opp.get("platform", opp.get("source_domain", "")),
            opp.get("type", "directory").replace("_", " ").title(),
            da,
            opp.get("relevance_score", 0),
            "Yes" if opp.get("is_dofollow", True) else "No",
            opp.get("status", "opportunity").replace("_", " ").title(),
            opp.get("source_url", opp.get("submission_url", "")),
            opp.get("notes", opp.get("ai_reasoning", "")),
            "Submit Now" if opp.get("status") == "opportunity" else opp.get("status", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [4,5,6] else "left")
            if col == 4:
                cell.fill = da_fill
                cell.font = Font(bold=True)
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
            if col == 10 and val == "Submit Now":
                cell.font = Font(color="1E40AF", bold=True)
        ws.row_dimensions[row].height = 35
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{row-1}"


# ─── Blog Ideas ───────────────────────────────────────────────────
def blog_ideas_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Blog Ideas")
    ideas = data.get("ideas", [])

    cols = ["#", "Blog Title", "Target Keyword", "Search Intent", "Priority Score", "Source", "AI Friendly", "Seasonal", "Content Gap", "AI Reasoning"]
    _title_row(ws, f"Blog Ideas — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 50, 30, 18, 14, 18, 12, 12, 12, 50]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, idea in enumerate(ideas, 1):
        score = idea.get("priority_score", 50)
        score_fill = PatternFill("solid", fgColor=("D1FAE5" if score >= 75 else ("FEF3C7" if score >= 50 else "FEE2E2")))
        values = [
            idx,
            idea.get("title", ""),
            idea.get("target_keyword", ""),
            idea.get("search_intent", "").replace("_", " ").title(),
            score,
            idea.get("source", "").replace("_", " ").title(),
            "✓" if idea.get("is_ai_friendly") else "",
            "✓" if idea.get("is_seasonal") else "",
            "✓" if idea.get("content_gap") else "",
            idea.get("ai_reasoning", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [5,7,8,9] else "left")
            if col == 5:
                cell.fill = score_fill
                cell.font = Font(bold=True)
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 40
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{row-1}"


# ─── Content Gap Analysis ─────────────────────────────────────────
def content_gap_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Content Gap Analysis")
    gaps = data.get("gaps", [])

    cols = ["#", "Topic / Keyword", "Gap Type", "Estimated Traffic", "Difficulty", "Priority", "Competitor Covering It", "Recommended Action", "Content Type"]
    _title_row(ws, f"Content Gap Analysis — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 40, 20, 18, 14, 12, 30, 45, 20]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, gap in enumerate(gaps, 1):
        priority = gap.get("priority", 50)
        values = [
            idx,
            gap.get("topic", gap.get("keyword", "")),
            gap.get("gap_type", "Content Missing"),
            gap.get("estimated_traffic", 0),
            gap.get("difficulty", "Medium").title(),
            priority,
            gap.get("competitor_covering_it", ""),
            gap.get("recommended_action", gap.get("our_opportunity", "")),
            gap.get("content_type", "Blog Post"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [4,5,6] else "left")
            if col == 6:
                cell.font = Font(bold=True, color=_score_color(int(priority or 50)))
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 35
        row += 1

    ws.freeze_panes = "A3"


# ─── SEO Tasks ────────────────────────────────────────────────────
def seo_tasks_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("SEO Tasks")
    tasks = data.get("tasks", [])

    cols = ["#", "Task Title", "Category", "Priority", "Status", "Estimated Impact", "Page URL", "Description", "Due Date", "AI Generated"]
    _title_row(ws, f"SEO Task List — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 45, 20, 12, 15, 16, 35, 50, 14, 14]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    priority_colors = {"critical": "FEE2E2", "high": "FEF3C7", "medium": "DBEAFE", "low": "D1FAE5"}
    for idx, task in enumerate(tasks, 1):
        priority = str(task.get("priority", "medium")).lower()
        p_fill = PatternFill("solid", fgColor=priority_colors.get(priority, "FFFFFF"))
        values = [
            idx,
            task.get("title", ""),
            task.get("category", "").replace("_", " ").title(),
            priority.upper(),
            task.get("status", "backlog").replace("_", " ").title(),
            task.get("estimated_impact", 0),
            task.get("page_url", ""),
            task.get("description", ""),
            str(task.get("due_date", ""))[:10] if task.get("due_date") else "",
            "🤖 AI" if task.get("ai_generated") else "Manual",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [4,5,6,9,10] else "left")
            if col == 4:
                cell.fill = p_fill
                cell.font = Font(bold=True)
            elif col == 6:
                cell.font = Font(bold=True, color=_score_color(int(val or 50)))
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 38
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{row-1}"


# ─── Competitor Analysis ──────────────────────────────────────────
def competitor_analysis_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Competitor Analysis")
    gaps = data.get("content_gaps", [])
    kw_gaps = data.get("keyword_gaps", [])

    cols = ["#", "Topic / Keyword", "Type", "Competitor", "Search Volume", "Difficulty", "Our Opportunity", "Priority", "Recommended Action"]
    _title_row(ws, f"Competitor Gap Analysis — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 40, 15, 28, 15, 14, 40, 10, 45]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    idx = 1
    for gap in gaps:
        values = [
            idx,
            gap.get("topic", ""),
            "Content Gap",
            gap.get("competitor_covering_it", ""),
            gap.get("estimated_traffic", 0),
            gap.get("difficulty", "Medium").title(),
            gap.get("our_opportunity", ""),
            gap.get("priority", 70),
            gap.get("content_type", "Create new page"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [5,6,8] else "left")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 35
        idx += 1
        row += 1

    for gap in kw_gaps:
        values = [
            idx,
            gap.get("keyword", ""),
            "Keyword Gap",
            gap.get("competitor_ranking", ""),
            gap.get("search_volume", 0),
            "Medium",
            "Rank for this keyword",
            75,
            gap.get("recommended_action", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if col in [5,6,8] else "left")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 35
        idx += 1
        row += 1

    ws.freeze_panes = "A3"


# ─── Internal Links ───────────────────────────────────────────────
def internal_links_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Internal Links")
    opps = data.get("opportunities", [])

    cols = ["#", "From Page", "To Page", "Suggested Anchor Text", "Reason", "Priority", "Status"]
    _title_row(ws, f"Internal Link Opportunities — {data.get('domain', '')} — {datetime.now():%Y-%m-%d}", len(cols))

    row = 2
    widths = [4, 45, 45, 30, 50, 12, 15]
    for i, col_name in enumerate(cols, 1):
        _header_style(ws, row, i, col_name, width=widths[i-1])
    ws.row_dimensions[row].height = 28
    row += 1

    for idx, opp in enumerate(opps, 1):
        values = [
            idx,
            opp.get("from_page", ""),
            opp.get("to_page", ""),
            opp.get("anchor_text", ""),
            opp.get("reason", ""),
            opp.get("priority", "medium").upper(),
            "Pending",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLORS["alt_row"])
        ws.row_dimensions[row].height = 35
        row += 1

    ws.freeze_panes = "A3"


# ─── Main Export Function ─────────────────────────────────────────
def generate_full_report(data: dict) -> bytes:
    """Generate a full multi-sheet Excel report."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    domain = data.get("domain", "website")

    # Cover / Summary sheet
    ws_cover = wb.create_sheet("📊 Summary", 0)
    ws_cover["A1"] = "SEO OS — Autonomous SEO Report"
    ws_cover["A1"].font = Font(bold=True, size=18, color=COLORS["header_bg"])
    ws_cover["A2"] = f"Domain: {domain}"
    ws_cover["A3"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"
    ws_cover["A4"] = f"SEO Score: {data.get('seo_score', 0)}/100"
    ws_cover["A4"].font = Font(bold=True, size=14, color=_score_color(data.get('seo_score', 0)))
    ws_cover["A5"] = f"Pages Crawled: {data.get('summary', {}).get('pages_crawled', 0)}"
    ws_cover["A6"] = f"Total Issues: {data.get('summary', {}).get('total_issues', 0)}"
    ws_cover["A7"] = f"Critical Issues: {data.get('summary', {}).get('critical_issues', 0)}"
    ws_cover["A8"] = f"Blog Ideas Generated: {len(data.get('ideas', []))}"
    ws_cover["A9"] = f"Backlink Opportunities: {len(data.get('opportunities', []))}"
    ws_cover["A10"] = f"SEO Tasks: {len(data.get('tasks', []))}"
    ws_cover.column_dimensions["A"].width = 40
    ws_cover.column_dimensions["B"].width = 30

    # Build all sheets
    if data.get("issues") or data.get("pages"):
        technical_audit_sheet(wb, data)

    if data.get("rankings"):
        rankings_sheet(wb, data)

    if data.get("opportunities"):
        backlinks_sheet(wb, data)

    if data.get("ideas"):
        blog_ideas_sheet(wb, data)

    if data.get("gaps"):
        content_gap_sheet(wb, data)

    if data.get("tasks"):
        seo_tasks_sheet(wb, data)

    if data.get("content_gaps") or data.get("keyword_gaps"):
        competitor_analysis_sheet(wb, data)

    if data.get("link_opportunities"):
        internal_links_sheet(wb, {"domain": domain, "opportunities": data.get("link_opportunities", [])})

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_sheet_only(sheet_type: str, data: dict) -> bytes:
    """Generate a single-sheet Excel file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if sheet_type == "technical_audit":
        technical_audit_sheet(wb, data)
    elif sheet_type == "rankings":
        rankings_sheet(wb, data)
    elif sheet_type == "backlinks":
        backlinks_sheet(wb, data)
    elif sheet_type == "blog_ideas":
        blog_ideas_sheet(wb, data)
    elif sheet_type == "content_gaps":
        content_gap_sheet(wb, data)
    elif sheet_type == "seo_tasks":
        seo_tasks_sheet(wb, data)
    elif sheet_type == "competitor":
        competitor_analysis_sheet(wb, data)
    elif sheet_type == "internal_links":
        internal_links_sheet(wb, data)
    else:
        ws = wb.create_sheet("Data")
        ws["A1"] = "No data"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
